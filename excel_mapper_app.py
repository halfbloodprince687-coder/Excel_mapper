import streamlit as st
from openpyxl import load_workbook
import re
from io import BytesIO
import traceback

try:
    # =========================
    # NORMALIZE TEXT
    # =========================
    def normalize(text):
        if text is None:
            return ""
        text = str(text).lower()
        text = re.sub(r"[^a-z]", "", text)
        return text

    def to_number(value):
        if value is None:
            return 0
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).replace(",", "").replace("₹", "").strip()
        match = re.search(r"-?\d+(\.\d+)?", text)
        if match:
            return float(match.group())
        return 0

    def write_to_cell(ws, cell_address, value):
        for merged_range in ws.merged_cells.ranges:
            if cell_address in merged_range:
                anchor = ws.cell(
                    row=merged_range.min_row,
                    column=merged_range.min_col
                )
                anchor.value = value
                return
        ws[cell_address].value = value

    # =========================
    # STREAMLIT APP
    # =========================
    st.set_page_config(page_title="Excel ETL Tool", page_icon="📊", layout="wide")
    st.title("📊 Excel ETL Tool")
    st.markdown("Upload your Excel files, configure settings, and run ETL.")

    # File uploads
    st.subheader("1️⃣ Upload Files")
    col1, col2, col3 = st.columns(3)
    with col1:
        source_file = st.file_uploader("Source Excel", type=["xlsx"], key="source")
    with col2:
        mapping_file = st.file_uploader("Mapping Excel", type=["xlsx"], key="mapping")
    with col3:
        target_file = st.file_uploader("Target Excel", type=["xlsx"], key="target")

    # Settings
    st.subheader("2️⃣ ETL Settings")
    factor = st.number_input("Multiplication Factor", value=1.0, step=1.0)

    # Sheet selection
    st.subheader("3️⃣ Select Sheet Names")
    source_sheet = target_sheet = None
    if source_file:
        source_wb_temp = load_workbook(source_file, read_only=True)
        source_sheet = st.selectbox("Source Sheet", source_wb_temp.sheetnames)
    if target_file:
        target_wb_temp = load_workbook(target_file, read_only=True)
        target_sheet = st.selectbox("Target Sheet", target_wb_temp.sheetnames)

    # Run ETL
    st.subheader("4️⃣ Run ETL")
    if st.button("Run ETL"):
        if not source_file or not mapping_file or not target_file:
            st.error("Please upload all three files!")
        elif not source_sheet or not target_sheet:
            st.error("Please select sheet names!")
        else:
            try:
                # Load workbooks
                source_wb = load_workbook(source_file, data_only=True)
                mapping_wb = load_workbook(mapping_file, data_only=True)
                target_wb = load_workbook(target_file)
                source_ws = source_wb[source_sheet]
                mapping_ws = mapping_wb["Sheet1"]
                target_ws = target_wb[target_sheet]

                # Load source data
                source_data = []
                for row in range(2, source_ws.max_row + 1):
                    key = source_ws[f"B{row}"].value
                    curr = source_ws[f"D{row}"].value
                    prev = source_ws[f"F{row}"].value
                    if key:
                        source_data.append({
                            "raw_key": str(key),
                            "norm_key": normalize(key),
                            "current": to_number(curr),
                            "previous": to_number(prev)
                        })

                st.info(f"Loaded {len(source_data)} source rows")

                # Process mapping
                filled = 0
                missing_keys = []
                for row in range(2, mapping_ws.max_row + 1):
                    map_key = mapping_ws[f"A{row}"].value
                    target_row = mapping_ws[f"B{row}"].value
                    if not map_key or not target_row:
                        continue
                    norm_map_key = normalize(map_key)
                    match = next((src for src in source_data if norm_map_key in src["norm_key"] or src["norm_key"] in norm_map_key), None)
                    if not match:
                        missing_keys.append(map_key)
                        continue
                    current_cell = f"G{int(target_row)}"
                    previous_cell = f"J{int(target_row)}"
                    curr_value = round(match["current"] * factor)
                    prev_value = round(match["previous"] * factor)
                    write_to_cell(target_ws, current_cell, curr_value)
                    write_to_cell(target_ws, previous_cell, prev_value)
                    filled += 1

                st.success(f"✅ Filled rows: {filled}")
                if missing_keys:
                    st.warning(f"⚠ Source keys not found for: {', '.join(missing_keys)}")

                # Save output
                output_buffer = BytesIO()
                target_wb.save(output_buffer)
                st.download_button(
                    label="📥 Download Filled Excel",
                    data=output_buffer.getvalue(),
                    file_name="target_filled.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            except Exception as e:
                st.error("❌ Unexpected error occurred. See error log.")
                with open("error_log.txt", "w") as f:
                    f.write(traceback.format_exc())

except Exception as e:
    print("Fatal error:", e)
    with open("fatal_error_log.txt", "w") as f:
        import traceback
        f.write(traceback.format_exc())
